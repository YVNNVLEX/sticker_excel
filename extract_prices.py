#!/usr/bin/env python3
import argparse
import sys
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

HEADER_ROW_FALLBACK = 7
COLUMN_LETTERS = {
    "article": "R",
    "ean": "T",
    "prix_club": "AM",
    "prix_public": "AN",
}


def normalize_header(text):
    if text is None:
        return ""
    text = str(text).lower()
    return "".join(ch for ch in text if ch.isalnum())


def find_header_row(ws, label, max_scan_rows=50):
    target = normalize_header(label)
    for idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1
    ):
        for cell in row:
            if isinstance(cell, str) and normalize_header(cell) == target:
                return idx
    return None


def normalize_text(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def normalize_ean(value):
    text = normalize_text(value)
    if not text:
        return None
    return text.replace(" ", "")


def format_price(value):
    if value is None:
        return "-"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def split_values(raw_groups):
    items = []
    for group in raw_groups:
        if group is None:
            continue
        if isinstance(group, str):
            group = [group]
        for item in group:
            for part in str(item).replace(";", ",").split(","):
                part = part.strip()
                if part:
                    items.append(part)
    return items


def build_index(ws):
    header_row = find_header_row(ws, "EAN Composant UVC") or HEADER_ROW_FALLBACK

    col_idx = {k: column_index_from_string(v) for k, v in COLUMN_LETTERS.items()}
    min_col = min(col_idx.values())
    max_col = max(col_idx.values())
    positions = {k: col_idx[k] - min_col for k in col_idx}

    ean_map = defaultdict(list)
    article_map = defaultdict(list)

    for row_idx, row in enumerate(
        ws.iter_rows(
            min_row=header_row + 1,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ),
        start=header_row + 1,
    ):
        article_val = row[positions["article"]]
        ean_val = row[positions["ean"]]
        prix_club_val = row[positions["prix_club"]]
        prix_public_val = row[positions["prix_public"]]

        if all(v is None for v in (article_val, ean_val, prix_club_val, prix_public_val)):
            continue

        article = normalize_text(article_val)
        ean = normalize_ean(ean_val)

        record = {
            "row": row_idx,
            "article": article or "",
            "ean": ean or "",
            "prix_club": prix_club_val,
            "prix_public": prix_public_val,
        }

        if ean:
            ean_map[ean].append(record)
        if article:
            article_map[article].append(record)

    return ean_map, article_map


def print_records(prefix, code, records):
    if not records:
        print(f"{prefix} {code} | NOT FOUND")
        return
    for record in records:
        print(
            f"{prefix} {code} | Article {record['article']} | EAN {record['ean']} "
            f"| PrixClub {format_price(record['prix_club'])} "
            f"| PrixGrandPublique {format_price(record['prix_public'])} "
            f"| Row {record['row']}"
        )


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Extract prix club and prix grand publique from an Excel file. "
            "If no --ean/--article is provided, an interactive menu is shown."
        )
    )
    parser.add_argument("excel_file", help="Path to the Excel file (.xlsx)")
    parser.add_argument(
        "--ean",
        nargs="+",
        action="append",
        default=[],
        help="EAN code(s). You can pass several values or repeat the option.",
    )
    parser.add_argument(
        "--article",
        nargs="+",
        action="append",
        default=[],
        help="Article reference(s). You can pass several values or repeat the option.",
    )

    args = parser.parse_args()

    ean_inputs = split_values(args.ean)
    article_inputs = split_values(args.article)

    try:
        wb = load_workbook(args.excel_file, read_only=True, data_only=True)
    except Exception as exc:
        print(f"Error opening workbook: {exc}", file=sys.stderr)
        return 1

    ws = wb.active
    ean_map, article_map = build_index(ws)

    if ean_inputs or article_inputs:
        for raw in ean_inputs:
            key = normalize_ean(raw)
            if key is None:
                print_records("EAN", raw, [])
                continue
            print_records("EAN", raw, ean_map.get(key, []))

        for raw in article_inputs:
            key = normalize_text(raw)
            if key is None:
                print_records("Article", raw, [])
                continue
            print_records("Article", raw, article_map.get(key, []))
        return 0

    while True:
        print("\nMenu:")
        print("1) Search by EAN")
        print("2) Search by Article")
        print("3) Reload file")
        print("4) Exit")
        choice = input("Choice: ").strip().lower()

        if choice in {"4", "exit", "quit", "q"}:
            break
        if choice == "3":
            try:
                wb = load_workbook(args.excel_file, read_only=True, data_only=True)
                ws = wb.active
                ean_map, article_map = build_index(ws)
                print("File reloaded.")
            except Exception as exc:
                print(f"Error opening workbook: {exc}")
            continue

        if choice == "1":
            raw = input("Enter EAN(s), separated by comma: ").strip()
            if not raw:
                continue
            for raw_code in split_values([raw]):
                key = normalize_ean(raw_code)
                if key is None:
                    print_records("EAN", raw_code, [])
                    continue
                print_records("EAN", raw_code, ean_map.get(key, []))
            continue

        if choice == "2":
            raw = input("Enter Article reference(s), separated by comma: ").strip()
            if not raw:
                continue
            for raw_code in split_values([raw]):
                key = normalize_text(raw_code)
                if key is None:
                    print_records("Article", raw_code, [])
                    continue
                print_records("Article", raw_code, article_map.get(key, []))
            continue

        print("Invalid choice.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
