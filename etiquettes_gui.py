#!/usr/bin/env python3
import argparse
import os
import sys
import time
import tkinter as tk
from collections import defaultdict
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from reportlab.graphics import renderPDF
from reportlab.graphics.barcode import createBarcodeDrawing
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

HEADER_ROW_FALLBACK = 7
COLUMN_LETTERS = {
    "article": "R",
    "ean": "T",
    "prix_club": "AM",
    "prix_public": "AN",
}


def normalize_header(text):
    if text is None:
        return ""
    text = str(text).lower()
    return "".join(ch for ch in text if ch.isalnum())


def find_header_row(ws, label, max_scan_rows=50):
    target = normalize_header(label)
    for idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1
    ):
        for cell in row:
            if isinstance(cell, str) and normalize_header(cell) == target:
                return idx
    return None


def normalize_text(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def normalize_ean(value):
    text = normalize_text(value)
    if not text:
        return None
    return text.replace(" ", "")


def format_price(value):
    if value is None:
        return "-"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def split_values(raw_groups):
    items = []
    for group in raw_groups:
        if group is None:
            continue
        if isinstance(group, str):
            group = [group]
        for item in group:
            for part in str(item).replace(";", ",").split(","):
                part = part.strip()
                if part:
                    items.append(part)
    return items


def build_index(ws):
    header_row = find_header_row(ws, "EAN Composant UVC") or HEADER_ROW_FALLBACK
    col_idx = {k: column_index_from_string(v) for k, v in COLUMN_LETTERS.items()}
    min_col = min(col_idx.values())
    max_col = max(col_idx.values())
    positions = {k: col_idx[k] - min_col for k in col_idx}

    ean_map = defaultdict(list)
    article_map = defaultdict(list)

    for row_idx, row in enumerate(
        ws.iter_rows(
            min_row=header_row + 1,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ),
        start=header_row + 1,
    ):
        article_val = row[positions["article"]]
        ean_val = row[positions["ean"]]
        prix_club_val = row[positions["prix_club"]]
        prix_public_val = row[positions["prix_public"]]

        if all(v is None for v in (article_val, ean_val, prix_club_val, prix_public_val)):
            continue

        article = normalize_text(article_val)
        ean = normalize_ean(ean_val)

        record = {
            "row": row_idx,
            "article": article or "",
            "ean": ean or "",
            "prix_club": prix_club_val,
            "prix_public": prix_public_val,
        }

        if ean:
            ean_map[ean].append(record)
        if article:
            article_map[article].append(record)

    return ean_map, article_map


def parse_label_size(text):
    if not text:
        return None
    cleaned = text.lower().replace("mm", "").strip()
    parts = cleaned.split("x")
    if len(parts) != 2:
        return None
    try:
        return float(parts[0]), float(parts[1])
    except ValueError:
        return None


def safe_filename(text):
    if not text:
        return "label"
    cleaned = []
    for ch in str(text):
        if ch.isalnum() or ch in {"-", "_"}:
            cleaned.append(ch)
        else:
            cleaned.append("_")
    result = "".join(cleaned).strip("_")
    return result or "label"


def create_barcode(value, width_pt, height_pt):
    value = value or ""
    if value.isdigit() and len(value) in {12, 13}:
        symbology = "EAN13"
    else:
        symbology = "Code128"
    try:
        drawing = createBarcodeDrawing(
            symbology, value=value, barHeight=height_pt, humanReadable=True
        )
    except Exception:
        drawing = createBarcodeDrawing(
            "Code128", value=value, barHeight=height_pt, humanReadable=True
        )
    if drawing.width and drawing.height:
        scale = min(width_pt / drawing.width, height_pt / drawing.height)
        drawing.scale(scale, scale)
    return drawing


def draw_label(c, record, width_pt, height_pt, margin_mm):
    margin = margin_mm * mm
    x = margin
    y = height_pt - margin
    line_gap = 2

    article = record.get("article", "") or "-"
    ean = record.get("ean", "") or "-"
    prix_club = format_price(record.get("prix_club"))
    prix_public = format_price(record.get("prix_public"))

    c.setFont("Helvetica-Bold", 7)
    c.drawString(x, y - 7, f"Article: {article}")
    y -= 7 + line_gap

    c.setFont("Helvetica-Bold", 12)
    c.setFillColorRGB(1, 0, 0)
    c.drawString(x, y - 12, f"Prix Club: {prix_club}")
    c.setFillColorRGB(0, 0, 0)
    y -= 12 + line_gap

    c.setFont("Helvetica", 7)
    c.drawString(x, y - 7, f"Prix Public: {prix_public}")
    y -= 7 + line_gap

    c.setFont("Helvetica", 6)
    c.drawString(x, y - 6, f"EAN: {ean}")
    y -= 6 + line_gap

    barcode_height = max(6 * mm, y - margin)
    barcode_width = max(1 * mm, width_pt - 2 * margin)
    barcode = create_barcode(ean, barcode_width, barcode_height)
    renderPDF.draw(barcode, c, x, margin)


def generate_pdf(records, output_path, label_width_mm, label_height_mm, margin_mm):
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    width_pt = label_width_mm * mm
    height_pt = label_height_mm * mm
    c = canvas.Canvas(output_path, pagesize=(width_pt, height_pt))
    for record in records:
        draw_label(c, record, width_pt, height_pt, margin_mm)
        c.showPage()
    c.save()


def print_pdf(path):
    if sys.platform.startswith("win"):
        try:
            os.startfile(path, "print")
            return True
        except Exception:
            return False
    return False


class LabelApp(tk.Tk):
    def __init__(self, excel_path, label_width_mm, label_height_mm, margin_mm, output_dir):
        super().__init__()
        self.title("Etiquettes UI")
        self.minsize(850, 600)

        self.excel_path = excel_path
        self.label_width_mm = label_width_mm
        self.label_height_mm = label_height_mm
        self.margin_mm = margin_mm
        self.output_dir = output_dir

        self.ean_map = {}
        self.article_map = {}
        self.results = []

        self.file_var = tk.StringVar(value=excel_path or "")
        self.status_var = tk.StringVar(value="Ready")
        self.search_var = tk.StringVar()
        self.search_type = tk.StringVar(value="EAN")
        self.detail_article = tk.StringVar(value="-")
        self.detail_ean = tk.StringVar(value="-")
        self.detail_prix_club = tk.StringVar(value="-")
        self.detail_prix_public = tk.StringVar(value="-")

        self._build_ui()

        if excel_path:
            self.load_file(excel_path)
        else:
            self.after(100, self.choose_file)

    def _build_ui(self):
        self.columnconfigure(0, weight=1)
        self.rowconfigure(2, weight=1)

        file_frame = ttk.Frame(self)
        file_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=6)
        file_frame.columnconfigure(1, weight=1)

        ttk.Label(file_frame, text="Fichier:").grid(row=0, column=0, sticky="w")
        ttk.Label(file_frame, textvariable=self.file_var).grid(
            row=0, column=1, sticky="ew", padx=(6, 6)
        )
        ttk.Button(file_frame, text="Choisir", command=self.choose_file).grid(
            row=0, column=2, padx=(0, 6)
        )
        ttk.Button(file_frame, text="Recharger", command=self.reload_file).grid(
            row=0, column=3
        )

        search_frame = ttk.Frame(self)
        search_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=6)
        search_frame.columnconfigure(2, weight=1)

        ttk.Label(search_frame, text="Recherche:").grid(row=0, column=0, sticky="w")
        search_type = ttk.Combobox(
            search_frame,
            textvariable=self.search_type,
            values=["EAN", "Article"],
            state="readonly",
            width=10,
        )
        search_type.grid(row=0, column=1, padx=(6, 6))
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var)
        search_entry.grid(row=0, column=2, sticky="ew")
        search_entry.bind("<Return>", lambda _event: self.run_search())
        ttk.Button(search_frame, text="Chercher", command=self.run_search).grid(
            row=0, column=3, padx=(6, 0)
        )
        ttk.Button(search_frame, text="Effacer", command=self.clear_search).grid(
            row=0, column=4, padx=(6, 0)
        )

        results_frame = ttk.Frame(self)
        results_frame.grid(row=2, column=0, sticky="nsew", padx=10, pady=6)
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)

        self.tree = ttk.Treeview(
            results_frame,
            columns=("article", "ean", "prix_club", "prix_public"),
            show="headings",
            selectmode="extended",
        )
        self.tree.heading("article", text="Article")
        self.tree.heading("ean", text="EAN")
        self.tree.heading("prix_club", text="Prix Club")
        self.tree.heading("prix_public", text="Prix Public")
        self.tree.column("article", width=200, anchor="w")
        self.tree.column("ean", width=160, anchor="w")
        self.tree.column("prix_club", width=100, anchor="center")
        self.tree.column("prix_public", width=120, anchor="center")
        self.tree.bind("<<TreeviewSelect>>", self.on_select)

        scroll = ttk.Scrollbar(results_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scroll.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        scroll.grid(row=0, column=1, sticky="ns")

        detail_frame = ttk.LabelFrame(self, text="Selection")
        detail_frame.grid(row=3, column=0, sticky="ew", padx=10, pady=6)
        detail_frame.columnconfigure(1, weight=1)

        ttk.Label(detail_frame, text="Article:").grid(row=0, column=0, sticky="w")
        ttk.Label(detail_frame, textvariable=self.detail_article).grid(
            row=0, column=1, sticky="w"
        )
        ttk.Label(detail_frame, text="EAN:").grid(row=1, column=0, sticky="w")
        ttk.Label(detail_frame, textvariable=self.detail_ean).grid(
            row=1, column=1, sticky="w"
        )
        ttk.Label(detail_frame, text="Prix Club:").grid(row=2, column=0, sticky="w")
        ttk.Label(
            detail_frame, textvariable=self.detail_prix_club, foreground="red"
        ).grid(row=2, column=1, sticky="w")
        ttk.Label(detail_frame, text="Prix Public:").grid(row=3, column=0, sticky="w")
        ttk.Label(detail_frame, textvariable=self.detail_prix_public).grid(
            row=3, column=1, sticky="w"
        )

        action_frame = ttk.Frame(self)
        action_frame.grid(row=4, column=0, sticky="ew", padx=10, pady=6)
        action_frame.columnconfigure(1, weight=1)

        ttk.Button(action_frame, text="Imprimer selection", command=self.print_selected).grid(
            row=0, column=0
        )
        ttk.Button(action_frame, text="Imprimer tous", command=self.print_all).grid(
            row=0, column=1, padx=(6, 0)
        )
        ttk.Label(
            action_frame,
            text=f"Format: {self.label_width_mm}x{self.label_height_mm} mm",
        ).grid(row=0, column=2, padx=(10, 0))
        ttk.Label(action_frame, textvariable=self.status_var).grid(
            row=0, column=3, sticky="e", padx=(10, 0)
        )

    def choose_file(self):
        path = filedialog.askopenfilename(
            title="Choisir un fichier Excel",
            filetypes=[("Excel", "*.xlsx")],
        )
        if path:
            self.load_file(path)

    def reload_file(self):
        if not self.excel_path:
            self.choose_file()
            return
        self.load_file(self.excel_path)

    def load_file(self, path):
        self.status_var.set("Chargement...")
        self.update_idletasks()
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            self.ean_map, self.article_map = build_index(ws)
            wb.close()
        except Exception as exc:
            messagebox.showerror("Erreur", f"Lecture impossible: {exc}")
            self.status_var.set("Erreur")
            return
        self.excel_path = path
        self.file_var.set(path)
        self.clear_results()
        self.status_var.set(
            f"Charges: {len(self.article_map)} articles, {len(self.ean_map)} ean"
        )

    def clear_results(self):
        self.results = []
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.update_details(None)

    def clear_search(self):
        self.search_var.set("")
        self.clear_results()

    def run_search(self):
        if not self.excel_path:
            messagebox.showinfo("Info", "Choisir un fichier Excel.")
            return
        raw = self.search_var.get().strip()
        if not raw:
            messagebox.showinfo("Info", "Entrer une valeur de recherche.")
            return
        search_terms = split_values([raw])
        records = []
        if self.search_type.get() == "EAN":
            for term in search_terms:
                key = normalize_ean(term)
                if not key:
                    continue
                records.extend(self.ean_map.get(key, []))
        else:
            for term in search_terms:
                key = normalize_text(term)
                if not key:
                    continue
                records.extend(self.article_map.get(key, []))

        unique = []
        seen = set()
        for record in records:
            key = (record["row"], record["article"], record["ean"])
            if key in seen:
                continue
            seen.add(key)
            unique.append(record)

        self.results = unique
        for item in self.tree.get_children():
            self.tree.delete(item)
        for idx, record in enumerate(unique):
            self.tree.insert(
                "",
                "end",
                iid=str(idx),
                values=(
                    record["article"],
                    record["ean"],
                    format_price(record["prix_club"]),
                    format_price(record["prix_public"]),
                ),
            )
        if not unique:
            messagebox.showinfo("Resultat", "Aucun resultat.")
        self.update_details(unique[0] if len(unique) == 1 else None)

    def on_select(self, _event):
        selected = self.tree.selection()
        if len(selected) != 1:
            self.update_details(None)
            return
        try:
            record = self.results[int(selected[0])]
        except (ValueError, IndexError):
            record = None
        self.update_details(record)

    def update_details(self, record):
        if not record:
            self.detail_article.set("-")
            self.detail_ean.set("-")
            self.detail_prix_club.set("-")
            self.detail_prix_public.set("-")
            return
        self.detail_article.set(record.get("article", "-") or "-")
        self.detail_ean.set(record.get("ean", "-") or "-")
        self.detail_prix_club.set(format_price(record.get("prix_club")))
        self.detail_prix_public.set(format_price(record.get("prix_public")))

    def get_selected_records(self):
        selected = self.tree.selection()
        records = []
        for item_id in selected:
            try:
                records.append(self.results[int(item_id)])
            except (ValueError, IndexError):
                continue
        return records

    def print_selected(self):
        records = self.get_selected_records()
        if not records:
            messagebox.showinfo("Info", "Selection vide.")
            return
        self._print_records(records)

    def print_all(self):
        if not self.results:
            messagebox.showinfo("Info", "Aucun resultat a imprimer.")
            return
        self._print_records(self.results)

    def _print_records(self, records):
        os.makedirs(self.output_dir, exist_ok=True)
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        if len(records) == 1:
            record = records[0]
            base = f"label_{safe_filename(record.get('article'))}_{safe_filename(record.get('ean'))}_{timestamp}.pdf"
        else:
            base = f"labels_{len(records)}_{timestamp}.pdf"
        output_path = os.path.join(self.output_dir, base)

        try:
            generate_pdf(
                records,
                output_path,
                self.label_width_mm,
                self.label_height_mm,
                self.margin_mm,
            )
        except Exception as exc:
            messagebox.showerror("Erreur", f"Generation PDF impossible: {exc}")
            return

        if not print_pdf(output_path):
            messagebox.showinfo(
                "Info", f"PDF genere. Impression automatique indisponible: {output_path}"
            )
            return
        self.status_var.set(f"Impression lancee: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="GUI pour rechercher des articles et imprimer des etiquettes."
    )
    parser.add_argument("excel_file", nargs="?", help="Fichier Excel (.xlsx)")
    parser.add_argument(
        "--label-size",
        default="50x25",
        help="Format etiquette en mm, ex: 50x25",
    )
    parser.add_argument("--label-width-mm", type=float, help="Largeur etiquette en mm")
    parser.add_argument("--label-height-mm", type=float, help="Hauteur etiquette en mm")
    parser.add_argument("--margin-mm", type=float, default=1.0, help="Marge en mm")
    parser.add_argument(
        "--output-dir",
        default="labels_output",
        help="Dossier de sortie pour les PDF",
    )
    args = parser.parse_args()

    width_mm = args.label_width_mm
    height_mm = args.label_height_mm
    if width_mm is None or height_mm is None:
        parsed = parse_label_size(args.label_size)
        if not parsed:
            print("Invalid --label-size, expected format like 50x25", file=sys.stderr)
            return 1
        width_mm, height_mm = parsed

    excel_path = args.excel_file
    if not excel_path and os.path.exists("ORCHESTRA_MARGE.xlsx"):
        excel_path = "ORCHESTRA_MARGE.xlsx"

    app = LabelApp(excel_path, width_mm, height_mm, args.margin_mm, args.output_dir)
    app.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
